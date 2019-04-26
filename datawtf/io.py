from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd

def loadExcel(file, **kwargs):
    '''
        This function is to load excel file as workbook
    '''
    return load_workbook(file, **kwargs)


def createExcel(**kwargs):
    '''
        This function is to create a new workbook
    '''
    return Workbook(**kwargs)


def getShape(workbook, **kwargs):
    '''
        get size of workbook
    '''
    return (workbook.max_row, workbook.max_column)


def indexToAlpha(numbers):
    '''
        Map index to excel column index
    '''
    if 0 < numbers < 27:
        return chr(numbers + 64)
    else:
        raise ValueError("Argument here must from 1 to 26")


def index(number):
    '''
        Map index to excel column index
        


        need fix!
    '''
    a = ''
    mod = number
    while(1):
        if(0 < mod < 27):
            return a + indexToAlpha(mod)
        
        else:
            number, mod = divmod(mod, 27)
            a = a + indexToAlpha(number)
            

def loadFile(filename):
    fileNameExtension = filename.split('.')[-1]
    if 'csv' in fileNameExtension:
        return pd.read_csv(filename)

    elif 'xls' in fileNameExtension:
        return pd.read_excel(filename)

    else:
        raise Exception("The file is not supported.!")


