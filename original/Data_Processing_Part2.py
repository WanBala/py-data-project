# -*- coding: utf-8 -*-
"""
Created on Fri Jan 19 09:42:18 2018

@author: YU-TING
"""

from openpyxl import load_workbook
from openpyxl import Workbook
import calendar
from copy import deepcopy

def month_create(R_Pointer,year,month,content,work_book,writing_book):
    if(month==4 or month==6 or month==9 or month==11):
        Day=30
    elif(month==2):
        if(calendar.isleap(year)):
            Day=29
        else:
            Day=28
    else:
        Day=31
    time=0
    content.append(year)
    content.append(month)
    for Cell in work_book['L'+str(R_Pointer):'AP'+str(R_Pointer)][0]:
        time+=1
        Cache=deepcopy(content)
        Cache.append(time)
        Cache.append(Cell.value)
        writing_book.append(Cache)
        if(time==Day):
            break

def Row_read(R_Pointer,work_book,writing_book):
    content=[]
    for cell in work_book['D'+str(R_Pointer):'I'+str(R_Pointer)][0]:
        content.append(cell.value)
    year=work_book['J'+str(R_Pointer)].value
    month=work_book['K'+str(R_Pointer)].value
    month_create(R_Pointer,year,month,content,work_book,writing_book)
#--------------Import Data------------------
Data=load_workbook("Try.xlsx")  #Import Data
Data_Sheet=Data.active
#一些有用的參數
Total_Rows=Data_Sheet.max_row
Total_Columns=Data_Sheet.max_column
#-----------------------------------------------
NewFile=Workbook()  
New_Sheet=NewFile.active
for R_Pointer in range(Total_Rows):   #R_Poiniter注意要從有資料的那行開始
    Row_read(R_Pointer+1,Data_Sheet,New_Sheet)
NewFile.save('balances.xlsx')     #Save Data






